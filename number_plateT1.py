import numpy as np
import cv2
import os
import easyocr
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import re
from collections import Counter
from ultralytics import YOLO
import threading

# Paths to directories and files
plates_dir = r"C:\Users\deven\Desktop\Number Plate\plates"
excel_path = r"C:\Users\deven\Desktop\Number Plate\Detected Plates.xlsx"
best_model_path = r'C:\Users\deven\Desktop\Number Plate\model\best.pt'
yolo_model_path = r'C:\Users\deven\Desktop\Number Plate\model\yolov8n.pt'

# Create directories if they do not exist
if not os.path.exists(plates_dir):
    os.makedirs(plates_dir)

# Create an Excel workbook and sheet if not exists
if not os.path.exists(excel_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Detected Plates"
    sheet.append(["Image Path", "Plate Number", "Date", "Time"])
else:
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

# Initialize EasyOCR reader
reader = easyocr.Reader(['en'], gpu=True)  # Enable GPU for faster OCR

# Load the YOLO models
best_model = YOLO(best_model_path)
yolo_model = YOLO(yolo_model_path)

# Function to preprocess image for better OCR results
def preprocess_image(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 11, 17, 17)
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return binary

# Function to validate the detected text
def is_valid_plate(text):
    # Generalized pattern to match different formats of number plates
    pattern = re.compile(r'^[A-Z0-9]{1,3}[-\s]?[A-Z0-9]{1,4}[-\s]?[A-Z0-9]{1,4}$')
    return bool(pattern.match(text.replace(" ", "").replace("-", "")))

# Function to get OCR results and find the most frequent valid text
def get_most_frequent_text(preprocessed_img):
    ocr_results = reader.readtext(preprocessed_img, detail=0, paragraph=True)
    valid_texts = [text.replace(" ", "").upper() for text in ocr_results if is_valid_plate(text)]
    if valid_texts:
        most_common_text = Counter(valid_texts).most_common(1)[0][0]
        return most_common_text
    return None

# Function to save extracted text and image path to Excel
def save_to_excel(img_path, plate_text):
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    date, time = current_datetime.split()
    sheet.append([img_path, plate_text, date, time])
    workbook.save(excel_path)
    print(f"Saved to Excel: {plate_text}")

# Load previously detected plates from Excel to ensure uniqueness
detected_plates = set()
for row in sheet.iter_rows(min_row=2, values_only=True):
    detected_plates.add(row[1])

# Frame capture and processing thread
class VideoProcessor(threading.Thread):
    def __init__(self):
        super().__init__()
        self.cap = cv2.VideoCapture(0)
        self.running = True
        self.count = 0
        self.skip_frames = 5  # Process every 5th frame

    def run(self):
        while self.running:
            ret, frame = self.cap.read()
            if not ret:
                break

            self.count += 1
            if self.count % self.skip_frames != 0:
                continue

            # Resize frame for faster processing
            frame = cv2.resize(frame, (640, 480))

            # Convert the image to RGB
            image_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

            # Perform detection with both models
            best_results = best_model.predict(image_rgb)
            yolo_results = yolo_model.predict(image_rgb)

            # Parse detection results from both models
            for results in [best_results, yolo_results]:
                for result in results:
                    for box in result.boxes:
                        xmin, ymin, xmax, ymax = map(int, box.xyxy[0])
                        confidence = box.conf[0]
                        class_id = box.cls[0]

                        # Extract the region of interest (ROI) containing the number plate
                        roi = frame[ymin:ymax, xmin:xmax]
                        preprocessed_img = preprocess_image(roi)

                        # Get the most frequent valid text from OCR results
                        plate_text = get_most_frequent_text(preprocessed_img)
                        if plate_text:
                            print(f"OCR Text: {plate_text}")

                            # Only save if it's a new plate text
                            if plate_text not in detected_plates:
                                detected_plates.add(plate_text)

                                # Draw the bounding box and extracted text on the frame
                                cv2.rectangle(frame, (xmin, ymin), (xmax, ymax), (0, 255, 0), 2)
                                cv2.putText(frame, plate_text, (xmin, ymax + 30), cv2.FONT_HERSHEY_COMPLEX_SMALL, 1, (255, 0, 0), 2)

                                # Save the ROI image
                                img_path = os.path.join(plates_dir, f"scanned_img_{self.count}.jpg")
                                cv2.imwrite(img_path, roi)
                                print(f"Image saved: {img_path}")

                                # Save the extracted text and metadata to the Excel file
                                save_to_excel(img_path, plate_text)

            cv2.imshow("Result", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):  # Exit if 'q' is pressed
                self.running = False
                break

    def stop(self):
        self.running = False
        self.cap.release()

# Start the video processing thread
processor = VideoProcessor()
processor.start()

# Wait for the processing to finish
processor.join()
cv2.destroyAllWindows()
